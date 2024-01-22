import sys, os
import pandas as pd
import numpy as np
import pickle
import random, string

from PyQt6.QtCore import *
from PyQt6.QtGui import *
from PyQt6.QtWidgets import *

import matplotlib
from matplotlib.backends.backend_qt import SubplotToolQt
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.lines import Line2D
from matplotlib.figure import Figure
from matplotlib.axes import Axes
matplotlib.use('QtAgg')

from JEMViewer2.file_handler import savefile, envs
from JEMViewer2.axeslinestool import BoolEdit, AliasButton
from JEMViewer2.basetoolbar import BaseToolbar

from JEMViewer2.deco_figure import DecoFigure
import JEMViewer2.stylesheet as ss
from matplotlib import colors as mcolors
import matplotlib.backends.qt_editor.figureoptions as figopt
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils.units import pixels_to_EMU

screen_dpi = 72

def randomname(n):
   randlst = [random.choice(string.ascii_lowercase) for i in range(n)]
   return ''.join(randlst)

types = ["table","plot","ndarray","customfunc"]
separators = ["space & tab", "tab", "comma", "customsep"]

class MyFigureCanvas(FigureCanvas):
    nd_pasted = pyqtSignal(str,np.ndarray)
    line_pasted = pyqtSignal(str,list)
    table_required = pyqtSignal(str,str)
    custom_loader = pyqtSignal(list)
    alias_pasted = pyqtSignal(str,Axes,bool)
    remove_required = pyqtSignal(int)
    def __init__(self, parent, toolbar, call_as_library, call_from, fig=None):
        if fig != None:
            self.fig = fig
            self.fig.set_dpi(screen_dpi)
        else:
            if call_as_library:
                self.fig = DecoFigure(call_from, dpi=screen_dpi)
            else:
                self.fig = Figure(dpi=screen_dpi)
            self.fig.add_subplot(111)
            # self.fig.axes[0].legend().set_draggable(True)
        self.call_as_library = call_as_library
        super().__init__(self.fig)
        self.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        self.setAcceptDrops(True)
        self.opacity = False
        self.setWindowFlags(Qt.Window)
        # self.setWindowFlags(self.windowFlags() | Qt.CustomizeWindowHint)
        # self.setWindowFlags(self.windowFlags() & ~Qt.WindowCloseButtonHint)
        self.keycid = self.mpl_connect('key_press_event',self._keypressevent)
        self.parent_ = parent
        self.mytoolbar = toolbar
        self.setFocusPolicy(Qt.StrongFocus)
        self.close_from_cui = False
        self.mdi = None

    def set_mdi(self, mdi):
        self.mdi = mdi

    def magnify(self, percent):
        base_logical_dpi = screen_dpi * self.device_pixel_ratio
        newratio = percent / 100
        currentratio = self.fig.get_dpi() / base_logical_dpi
        self.fig.set_dpi(base_logical_dpi * newratio)
        if self.mdi == None:
            canvassize = self.size()
            w = int(canvassize.width() * newratio / currentratio)
            h = int(canvassize.height() * newratio / currentratio)
            self.resize(w,h)
        else:
            canvassize = self.size()
            mdisize = self.mdi.size()
            dw, dh = mdisize.width() - canvassize.width(), mdisize.height() - canvassize.height()
            w = int(canvassize.width() * newratio / currentratio) + dw
            h = int(canvassize.height() * newratio / currentratio) + dh
            self.mdi.resize(w,h)

    def set_window_title(self, id, prefix=None):
        if prefix != None:
            self.prefix = prefix
        self.fig_id = id
        if self.call_as_library:
            self.fig.set_id(id)
        self.setWindowTitle(f"{self.prefix} | Figure:{id}")

    def dragEnterEvent(self,event):
        event.accept()

    def data_from_clipboard(self):
        clipboard = QApplication.clipboard()
        lines = clipboard.text().splitlines()
        data = []
        for line in lines:
            data.append(line.split())
        return np.array(data,float).T

    def _keypressevent(self, event):
        if event.key in ['ctrl+v', 'cmd+v']:
            if event.inaxes == None:
                return
            new_data = self.data_from_clipboard()
            ax = event.inaxes
            self._add_newplot(new_data, ax)
        elif event.key in ['ctrl+n','cmd+n']:
            new_data = self.data_from_clipboard()
            self._add_newndarray(new_data)
        elif event.key in ['ctrl+t','cmd+t']:
            new_data = QApplication.clipboard().text()
            self.table_required.emit(new_data,DDHandler.delimiters[DDHandler.separator])
        elif event.key == 't':
            self.opacity = not self.opacity
            value = 0.5 if self.opacity else 1.0
            self.setWindowOpacity(value)
        elif event.key == 'p':
            self.parent_.raise_()
        #elif event.key in ['ctrl+t', 'cmd+t']:
        #    self.table_required.emit()

    def _add_newplot(self,data,ax,label=None):
        new_name = randomname(4)
        lines = []
        figaxid = self._identify_figaxid(ax)
        if data.shape[0] > 1:
            for i in range(1,len(data)):
                line, = ax.plot(data[0],data[i],label=label)
                lines.append(line)
            self.line_pasted.emit(new_name,lines)
            savefile.save_plot(new_name, figaxid, data, label)
        self.draw()
    
    def _identify_figaxid(self, ax):
        return {
            "figs": self.fig_id,
            "axes": self.fig.axes.index(ax)
        }

    def _add_newndarray(self,data):
        new_name = randomname(4)
        savefile.save_npdata(new_name,data)
        self.nd_pasted.emit(new_name,data)

    def _open_newplot(self,filename,inaxes):
        new_data = np.loadtxt(filename,delimiter=DDHandler.delimiters[DDHandler.separator]).T
        self._add_newplot(new_data,inaxes,label=os.path.basename(filename))

    def _open_newndarray(self,filename):
        new_data = np.loadtxt(filename,delimiter=DDHandler.delimiters[DDHandler.separator]).T
        self._add_newndarray(new_data)

    def dropEvent(self,event):
        event.accept()
        if type(event.source()) == AliasButton:
            alias = event.source().text()
            mod = event.modifiers()
            inaxes = self.inaxes(self.mouseEventCoords(event.position()))
            if inaxes == None: return
            self._move_plot(alias, inaxes, mod)
            return
        if event.mimeData().hasUrls():
            event.accept()
            for url in event.mimeData().urls():
                file = url.toLocalFile()
                if DDHandler.type == "table":
                    self.table_required.emit(file,DDHandler.delimiters[DDHandler.separator])
                elif DDHandler.type == "plot":
                    inaxes = self.inaxes(self.mouseEventCoords(event.position()))
                    if inaxes == None: return
                    self._open_newplot(file,inaxes)
                elif DDHandler.type == "ndarray":
                    self._open_newndarray(file)
                elif DDHandler.type == "customfunc":
                    inaxes = self.inaxes(self.mouseEventCoords(event.position()))
                    if inaxes == None: return
                    figaxid = self._identify_figaxid(inaxes)
                    lis = [DDHandler.functionname, file, inaxes, randomname(4), figaxid]
                    self.custom_loader.emit(lis)
                
    def focusInEvent(self, a0: QFocusEvent) -> None:
        self.mytoolbar.focused_canvas = self
        return super().focusInEvent(a0)

    def _move_plot(self, alias, ax, mod):
        self.alias_pasted.emit(alias, ax, mod.name == 'ControlModifier')

    def close_(self):
        self.close_from_cui = True
        self.mpl_disconnect(self.keycid)
        self.close()
        
    def closeEvent(self, e):
        if self.close_from_cui:
            e.accept()
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Remove")
            msg.setText("Are you sure to remove this figure?\n or close JEM viewer?")
            msg.setIcon(QMessageBox.Icon.Question)
            rem = msg.addButton("Remove figure", QMessageBox.ButtonRole.AcceptRole)
            cancel = msg.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
            close = msg.addButton("Close JEM", QMessageBox.ButtonRole.ActionRole)
            msg.setDefaultButton(cancel)
            msg.exec()

            if msg.clickedButton() == rem:
                self.remove_required.emit(self.fig_id)
                savefile.save_removefigure(self.fig_id)
            elif msg.clickedButton() == close:
                self.parent_.close()
            e.ignore()


class DDHandler(QDialog):
    type = "table"
    separator = "space & tab"
    functionname = ""
    delimiters = {
        "space & tab": None,
        "tab": "\t",
        "comma": ",",
        "customsep": "",
    }
    def __init__(self,parent):
        super().__init__(parent)
        self.setWindowTitle("Drag and Drop settings")
        self._create_main_frame()

    def _create_main_frame(self):
        self.btns = {}
        layout = QHBoxLayout()
        self.typeG = QGroupBox()
        self.typeG.setTitle("Open type")
        inner_type = QVBoxLayout()
        for type in types:
            btn = QRadioButton()
            btn.setFocusPolicy(Qt.NoFocus)
            tmp = QHBoxLayout()
            tmp.setAlignment(Qt.AlignLeft)
            tmp.addWidget(btn)
            if type == DDHandler.type:
                btn.setChecked(True)
            if type == 'customfunc':
                self.loaderbox = QLineEdit()
                self.loaderbox.setFixedWidth(80)
                self.loaderbox.setText(DDHandler.functionname)
                tmp.addWidget(self.loaderbox)         
                inner_type.addLayout(tmp)       
            else:
                tmp.addWidget(QLabel(type))
                inner_type.addLayout(tmp)
            self.btns[type] = btn
        self.typeG.setLayout(inner_type)
        self.separatorG = QGroupBox()
        self.separatorG.setTitle("Separator")
        inner_sep = QVBoxLayout()
        for separator in separators:
            btn = QRadioButton()
            btn.setFocusPolicy(Qt.NoFocus)
            tmp = QHBoxLayout()
            tmp.setAlignment(Qt.AlignLeft)
            tmp.addWidget(btn)
            if separator == DDHandler.separator:
                btn.setChecked(True)
            if separator == 'customsep':
                self.separatorbox = QLineEdit()
                self.separatorbox.setFixedWidth(80)
                self.separatorbox.setText(DDHandler.delimiters['customsep'])
                tmp.addWidget(self.separatorbox)         
                inner_sep.addLayout(tmp)       
            else:
                tmp.addWidget(QLabel(separator))
                inner_sep.addLayout(tmp)
            self.btns[separator] = btn
        self.separatorG.setLayout(inner_sep)
        layout.addWidget(self.typeG)
        layout.addWidget(self.separatorG)
        self.setLayout(layout)

    def update(self):
        for type in types:
            if self.btns[type].isChecked():
                DDHandler.type = type
        for seperator in separators:
            if self.btns[seperator].isChecked():
                DDHandler.separator = seperator
        DDHandler.functionname = self.loaderbox.text()
        DDHandler.delimiters['customsep'] = self.separatorbox.text()

    @staticmethod
    def set_loader(loader,separator=0):
        """setting loader

        Args:
            loader (str): loader name choose from {table, plot, ndarray} or input custom function name
            separator (int or str, optional): separator. choose from {0:space & tab, 1:tab, 2:comma} or input custom delimiter. Defaults to 0.
        """
        if loader in ["table", "plot", "ndarray"]:
            DDHandler.type = loader
        else:
            DDHandler.type = 'customfunc'
            DDHandler.functionname = loader
        if type(separator) == int:
            DDHandler.separator = separators[separator]
        elif separator in ["space & tab", "tab", "comma"]:
            DDHandler.separator = separator
        else:
            DDHandler.separator = 'customsep'
            DDHandler.delimiters['customsep'] = separator

    @staticmethod
    def show(parent):
        dialog = DDHandler(parent)
        dialog.exec_()
        dialog.update()


class MyToolbar(BaseToolbar):
    def __init__(self, parent, is_floatmode = True):
        self.mpl_toolbars = {}
        self.parent = parent
        self.tmp_ext = 'jpg'
        base = (
                ('Loader', 'Set loader type', os.path.join(envs.RES_DIR,'dd'), 'loader', None, False),
                ('AddFigure', 'Add figure', os.path.join(envs.RES_DIR,'addfigure'), 'addfigure', None, False),
                # ("EnableLineDrag", "Enable line drag", os.path.join(envs.RES_DIR,'linedrag'), 'enable_line_drag', None, True),
                (None, None, None, None, None, False),
        )
        if is_floatmode:
            mode = (
                    ('SwitchMode', 'Switch mode', os.path.join(envs.RES_DIR,'switchmode'), 'switch_mode', None, False),
                    ('Popup', 'Popup figures', os.path.join(envs.RES_DIR,'popup'), 'popup', None, False),
                    (None, None, None, None, None, False),
            )
        else:
            mode = (
                    ('SwitchMode', 'Switch mode', os.path.join(envs.RES_DIR,'switchmode'), 'switch_mode', None, False),
                    ('Tiling', 'Tiling figures', os.path.join(envs.RES_DIR,'tile'), 'tiling', None, False),
                    ('Cascading', 'Cascading figures', os.path.join(envs.RES_DIR,'cascade'), 'cascading', None, False),
                    (None, None, None, None, None, False),
            )
        tool = (
                ('AxesTool', 'Show AxesTool', os.path.join(envs.RES_DIR,'axestool'), 'axestool', None, False),
                ('LinesTool', 'Show LinesTool', os.path.join(envs.RES_DIR,'linestool'), 'linestool', None, False),
                ('TextsTool', 'Show TextsTool', os.path.join(envs.RES_DIR,'textstool'), 'textstool', None, False),
                (None, None, None, None, None, False),
        )
        mpl = (
                ('Home', 'Reset original view', 'home', 'home', None, False),
                ('Back', 'Back to previous view', 'back', 'back', None, False),
                ('Forward', 'Forward to next view', 'forward', 'forward', None, False),
                (None, None, None, None, None, False),
                ('Pan',
                'Left button pans, Right button zooms\n'
                'x/y fixes axis, CTRL fixes aspect',
                'move', 'pan', None, True),
                ('Zoom', 'Zoom to rectangle\nx/y fixes axis', 'zoom_to_rect', 'zoom', None, True),
                ('Subplots', 'Configure subplots', 'subplots', 'configure_subplots', None, False),
                (None, None, None, None, None, False),
        )
        save = (
                ('Save', 'Save the figure', 'filesave', 'save_figure', None, False),
                ('SaveAnim', 'Save figures for ppt animation', os.path.join(envs.RES_DIR,'savefiganim'), 'save_figure_for_animation', None, False),
                ('CopyFig', 'Copy figure to clipboard', os.path.join(envs.RES_DIR,'clipboard'), 'copy_figure_to_clipboard', 'clipboard_menu', False),
                ('Excel', 'Export data to excel', os.path.join(envs.RES_DIR,'excel'), 'export_data_to_excel', None, False),

        )
        self.toolitems = base + mode + tool + mpl + save
        super().__init__(parent)
        self.loc_label = self.parent.bar

    def loader(self):
        DDHandler.show(self)

    def popup(self):
        self.parent.raise_figure_widgets()

    def tiling(self):
        self.parent.tiling()

    def cascading(self):
        self.parent.cascading()

    def switch_mode(self):
        self.parent.switch_mode()

    def enable_line_drag(self):
        self.parent.linestool.set_line_draggable(self.actions['enable_line_drag'].isChecked())

    def linestool(self):
        self.parent.show_linestool()

    def axestool(self):
        self.parent.show_axestool()

    def textstool(self):
        self.parent.show_textstool()

    def addfigure(self):
        self.parent.add_figure()
        savefile.save_addfigure()

    def back(self):
        self.pop_message()    
        self.mpl_toolbars[self.focused_canvas].back()

    def forward(self):
        if self.focused_canvas == None:
            self.pop_message()
            return
        self.mpl_toolbars[self.focused_canvas].forward()

    def configure_subplots(self):
        if self.focused_canvas == None:
            self.pop_message()
            return
        if not MySubplotToolQt.active:
            self._subplot_dialog = MySubplotToolQt(self.focused_canvas, self)
        self._subplot_dialog.show()
        self._subplot_dialog.raise_()

    def save_figure(self):
        if self.focused_canvas == None:
            self.pop_message()
            return
        self.mpl_toolbars[self.focused_canvas].save_figure()

    def save_figure_for_animation(self):
        if self.focused_canvas == None:
            self.pop_message()
            return
        if not SaveForAnimationDialog.active:
            self._savefiganim_dialog = SaveForAnimationDialog(self.focused_canvas, self)
        self._savefiganim_dialog.show()
        self._savefiganim_dialog.raise_()

    def copy_figure_to_clipboard(self):
        if self.focused_canvas == None:
            self.pop_message()
            return
        figpath = os.path.join(envs.TEMP_DIR,'temp.' + self.tmp_ext)
        self.focused_canvas.figure.savefig(figpath)
        clipboard = QApplication.clipboard()
        clipboard.setPixmap(QPixmap(figpath))
        #os.remove(figpath)

    def export_data_to_excel(self):
        marker_map = {
            'point': "dot",
            'pixel': "dot",
            'circle': "circle",
            'triangle_down': "triangle",
            'triangle_up': "triangle",
            'triangle_left': "triangle",
            'triangle_right': "triangle",
            'tri_down': "triangle",
            'tri_up': "triangle",
            'tri_left': "triangle",
            'tri_right': "triangle",
            'octagon': "square",
            'square': "square",
            'pentagon': "square",
            'star': "star",
            'hexagon1': "square",
            'hexagon2': "square",
            'plus': "x",
            'x': "x",
            'diamond':"diamond",
            'thin_diamond':"diamond",
            'vline': "dash",
            'hline': "dash",
            'plus_filled':"x",
            'x_filled':"x",
            'tickleft': "triangle",
            'tickright': "triangle",
            'tickup': "triangle",
            'tickdown': "triangle",
            'caretleft': "triangle",
            'caretright': "triangle",
            'caretup': "triangle",
            'caretdown': "triangle",
            'caretleftbase': "triangle",
            'caretrightbase': "triangle",
            'caretupbase': "triangle",
            'caretdownbase': "triangle",
            '': 'nothing'
        }
        line_map = {
            "Solid": "solid",
            "Dashed": "dash",
            "Dashdot": "sysDashDot",
            "Dotted": "dot",
            "None": "None"
        }
        if self.focused_canvas == None:
            self.pop_message()
            return
        startpath = os.path.expanduser(matplotlib.rcParams['savefig.directory'])
        basename = os.path.basename(self.focused_canvas.get_default_filename())
        base, ext = os.path.splitext(basename)
        start = os.path.join(startpath, base)
        filepath = QFileDialog.getSaveFileName(self,"Project name",start,"Excel Book (*.xlsx)")
        if filepath[0] == '':
            return
        filepath = filepath[0]
        workbook = Workbook()
        for i, fig in enumerate(self.parent.figs):
            for j, ax in enumerate(fig.axes):
                sheet = workbook.create_sheet(f"fig{i}ax{j}", 0)
                chart = ScatterChart()
                chart.x_axis.title = ax.get_xlabel()
                chart.y_axis.title = ax.get_ylabel()
                chart.x_axis.scaling.min = ax.get_xlim()[0]
                chart.x_axis.scaling.max = ax.get_xlim()[1]
                chart.y_axis.scaling.min = ax.get_ylim()[0]
                chart.y_axis.scaling.max = ax.get_ylim()[1]
                if ax.get_xscale() == 'log':
                    chart.x_axis.scaling.logBase = 10
                if ax.get_yscale() == 'log':
                    chart.y_axis.scaling.logBase = 10
                chart.title = ax.get_title()
                chart.height = 15
                chart.width = 15
                column = 1
                for line in ax.lines:
                    if not line.get_visible():
                        continue
                    x = line.get_xdata()
                    y = line.get_ydata()
                    data_length = len(x)
                    sheet.cell(row=1, column=column, value=line.get_label())
                    sheet.cell(row=2, column=column, value=ax.get_xlabel())
                    sheet.cell(row=2, column=column+1, value=ax.get_ylabel())
                    for i in range(data_length):
                        sheet.cell(row=i+3, column=column, value=x[i])
                        sheet.cell(row=i+3, column=column+1, value=y[i])
                    x_wl = Reference(sheet, min_col=column, min_row=3, max_row=data_length+2)
                    y_wl = Reference(sheet, min_col=column+1, min_row=3, max_row=data_length+2)
                    series = Series(y_wl, x_wl, title=line.get_label())
                    linestyle = figopt.LINESTYLES[line.get_linestyle()]
                    if linestyle != 'None':
                        series.graphicalProperties.line.dashStyle = line_map[linestyle]
                        color = mcolors.to_hex(mcolors.to_rgb(line.get_color())).replace('#','')
                        series.graphicalProperties.line.solidFill = color
                        series.graphicalProperties.line.width = pixels_to_EMU(line.get_linewidth())
                    else:
                        series.graphicalProperties.line.noFill = True
                    marker = figopt.MARKERS[line.get_marker()]
                    if marker != 'nothing':
                        series.marker.symbol = marker_map[marker]
                        series.marker.size = line.get_markersize()
                        incolor = mcolors.to_hex(mcolors.to_rgb(line.get_markerfacecolor())).replace('#','')
                        series.marker.graphicalProperties.solidFill = incolor
                        outcolor = mcolors.to_hex(mcolors.to_rgb(line.get_markeredgecolor())).replace('#','')
                        series.marker.graphicalProperties.line.solidFill = outcolor
                        series.marker.graphicalProperties.line.width = pixels_to_EMU(line.get_markeredgewidth())
                    chart.series.append(series)
                    column += 2
                sheet.add_chart(chart,"A10")
        workbook.save(filepath)

    def clipboard_menu(self):
        def set_ext():
            sender = self.sender()
            self.tmp_ext = sender.text()
        menu = QMenu(self)
        for ext in ['png', 'jpg', 'pdf', 'svg']:
            action = menu.addAction(ext)
            action.setCheckable(True)
            action.setChecked(ext == self.tmp_ext)
            action.triggered.connect(set_ext)
        menu.exec(QCursor.pos())

    def home(self):
        if self.focused_canvas == None:
            self.pop_message()
            return
        self.mpl_toolbars[self.focused_canvas].home()

    def zoom(self):
        for toolbar in self.mpl_toolbars.values():
            toolbar.zoom()
        self.update_buttons_checked()

    def pan(self):
        for toolbar in self.mpl_toolbars.values():
            toolbar.pan()
        self.update_buttons_checked()

    def update_buttons_checked(self):
        if len(self.mpl_toolbars) > 0:
            self.actions['pan'].setChecked(list(self.mpl_toolbars.values())[0].mode.name == 'PAN')
            self.actions['zoom'].setChecked(list(self.mpl_toolbars.values())[0].mode.name == 'ZOOM')

    def add_canvas(self, canvas):
        if self.actions['pan'].isChecked(): self.actions['pan'].trigger()
        if self.actions['zoom'].isChecked(): self.actions['zoom'].trigger()
        toolbar = NavigationToolbar(canvas, None)
        self.mpl_toolbars[canvas] = toolbar
        toolbar.locLabel = self.loc_label
        self.focused_canvas = canvas

    def remove_canvas(self, canvas):
        self.mpl_toolbars.pop(canvas)
        self.focused_canvas = None

    def pop_message(self, message="Please focus a figure widget."):
        QMessageBox.critical(self,"Error",message)


class MySubplotToolQt(SubplotToolQt):
    active = False
    def __init__(self, canvas, parent):
        MySubplotToolQt.active = True
        self.fig_id = canvas.fig_id
        tight = canvas.fig.get_tight_layout()
        super().__init__(canvas.fig, parent)
        self.setWindowTitle(f"Parameters for {canvas.windowTitle()}")
        canvas.fig.set_tight_layout(tight)

    def _on_value_changed(self):
        self._figure.set_tight_layout(False)
        return super()._on_value_changed()

    def _tight_layout(self):
        self._figure.set_tight_layout(True)
        return super()._tight_layout()

    def closeEvent(self, event):
        MySubplotToolQt.active = False
        parameter = ",".join(f"{attr}={spinbox.value():.3}" for attr, spinbox in self._spinboxes.items())
        savefile.save_subplotsparam(self._figure.get_tight_layout(), self.fig_id, parameter)
        event.accept()


class SaveForAnimationDialog(QDialog):
    active = False
    def __init__(self,figure_canvas,parent):
        super().__init__(parent)
        self.setWindowTitle("Save figures for ppt animation")
        self.figure = figure_canvas.fig
        self.fig_id = figure_canvas.fig_id
        self.table = QTableWidget(self)
        layout = QVBoxLayout()
        blayout = QHBoxLayout()
        self.add_btn = QPushButton("Add scene")
        self.del_btn = QPushButton("Del sence")
        self.save_btn = QPushButton("Save")
        blayout.addWidget(self.add_btn)
        blayout.addWidget(self.del_btn)
        blayout.addWidget(self.save_btn)
        self.add_btn.clicked.connect(self.add)
        self.del_btn.clicked.connect(self.remove)
        self.save_btn.clicked.connect(self.save)
        self.setLayout(layout)
        layout.addWidget(self.table)
        layout.addLayout(blayout)
        self.make_alias()
        self.header = ["frame"] + list(self.lines.keys())
        self.table.setRowCount(len(self.header))
        self.table.setVerticalHeaderLabels(self.header)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContentsOnFirstShow)
        self.initialize(len(self.lines))

    def make_alias(self):
        self.lines = {}
        for i, ax in enumerate(self.figure.axes):
            for j, l in enumerate(ax.lines):
                alias = f"fig{self.fig_id}ax{i}l{j}"
                self.lines[alias] = l
    
    def initialize(self, initial):
        self.table.setColumnCount(0)
        for n in range(initial):
            self.add_widget(n)

    def add_widget(self, i):
        self.table.insertColumn(i)
        for j in range(len(self.header)):
            self.table.setCellWidget(j,i,BoolEdit(initial=(i+1==j or i+j==0)))

    def add(self):
        self.add_widget(self.table.columnCount())

    def remove(self):
        if self.table.columnCount() > 1:
            self.table.setColumnCount(self.table.columnCount()-1)

    def result(self):
        scene = []
        for i in range(self.table.columnCount()):
            visible = {}
            for j,l in enumerate(self.header):
                visible[l] = self.table.cellWidget(j,i).value()
            scene.append(visible)
        return scene

    def save(self):
        filetypes = self.figure.canvas.get_supported_filetypes_grouped()
        sorted_filetypes = sorted(filetypes.items())
        default_filetype = self.figure.canvas.get_default_filetype()

        startpath = os.path.expanduser(matplotlib.rcParams['savefig.directory'])
        start = os.path.join(startpath, self.figure.canvas.get_default_filename())
        filters = []
        selectedFilter = None
        for name, exts in sorted_filetypes:
            exts_list = " ".join(['*.%s' % ext for ext in exts])
            filter = '%s (%s)' % (name, exts_list)
            if default_filetype in exts:
                selectedFilter = filter
            filters.append(filter)
        filters = ';;'.join(filters)

        fname, filter = QFileDialog.getSaveFileName(
            self.figure.canvas.parent(), "Choose a filename to save to", start,
            filters, selectedFilter)
        if fname:
            # Save dir for next time, unless empty str (i.e., use cwd).
            if startpath != "":
                matplotlib.rcParams['savefig.directory'] = os.path.dirname(fname)
            try:
                self.save_animation(fname)
            except Exception as e:
                QMessageBox.critical(self, "Error saving file", str(e), QMessageBox.StandardButton.Ok)

    def save_animation(self,fname):
        prefix, ext = os.path.splitext(fname)
        scene = self.result()
        initial_visible = {h:self.lines[h].get_visible() for h in self.header[1:]}
        initial_framecolor = self.figure.axes[0].xaxis.label.get_color()
        def set_framecolor(color):
            for axis in self.figure.axes:
                axis.xaxis.label.set_color(color)
                axis.yaxis.label.set_color(color)
                axis.tick_params(colors=color,which='both')
                for i in ['left','right','top','bottom']:
                    axis.spines[i].set_color(color)
        def draw_frame(visible):
            if visible:
                set_framecolor(initial_framecolor)
            else:
                set_framecolor("none")
        for i, s in enumerate(scene):
            for key, vis in s.items():
                if key == "frame":
                    draw_frame(vis)
                else:
                    self.lines[key].set_visible(vis)
            self.figure.savefig('{0}{1:02d}{2}'.format(prefix,i,ext))
        set_framecolor(initial_framecolor)
        for key, vis in initial_visible.items():
            self.lines[key].set_visible(vis)
        self.figure.canvas.draw()

    def closeEvent(self, event):
        SaveForAnimationDialog.active = False
        event.accept()

